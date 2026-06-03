"""
Wine Cellar — one-time import script
Reads from: IG BBR FR Wine Personal Post Handover NEW 2026 05 17.xlsx
    - "New Consoliated" tab  → UK in-bond wines
    - "VINOTHEQUE" tab       → Vienna professional storage
    - "Wien" tab             → Vienna home cellar
Inserts into: Supabase `wines` table (and `sell_listings` for listed wines)

Run once:
    pip install -r requirements.txt
    cp .env.example .env   # fill in your Supabase credentials
    python import.py
"""

import os
import re
import openpyxl
from dotenv import load_dotenv
from supabase import create_client

load_dotenv()

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_KEY"]
EXCEL_FILE   = "IG BBR FR Wine Personal Post Handover NEW 2026 05 17.xlsx"

supabase = create_client(SUPABASE_URL, SUPABASE_KEY)


# ── helpers ──────────────────────────────────────────────────────────────────

def safe_vintage(v):
    """Parse vintage — returns int or None. 'NV', blanks etc. → None."""
    if v is None:
        return None
    try:
        return int(float(str(v).strip()))
    except Exception:
        return None  # NV, blanks, anything non-numeric

def safe_float(v):
    if v is None:
        return None
    try:
        return float(str(v).replace(",", ".").strip())
    except Exception:
        return None

def parse_bottle_count(v):
    """'6', '12', 6, 12  →  integer (always plain numbers in this sheet)"""
    if v is None:
        return None
    try:
        return int(float(str(v).strip()))
    except Exception:
        return None

def cost_per_bottle(cost_total, bottle_count):
    if cost_total and bottle_count:
        return round(cost_total / bottle_count, 2)
    return None

MERCHANT_STORAGE = {
    "BBR":           "BBR Bond",
    "Justerini":     "Justerini Bond",
    "IG":            "IG Bond",
    "Hatton Edwards":"Hatton Edwards",
    "Fine and Rare": "Fine and Rare",
    "Lay Wheeler":   "Lay Wheeler",
}

MERCHANT_ALIASES = {
    "Fine + Rare":   "Fine and Rare",
    "Fine & Rare":   "Fine and Rare",
    "HE":            "Hatton Edwards",
}

def normalise_merchant(raw):
    if not raw:
        return None
    s = str(raw).strip()
    return MERCHANT_ALIASES.get(s, s)

def storage_for_merchant(merchant):
    return MERCHANT_STORAGE.get(merchant, f"{merchant} Bond" if merchant else "Unknown")


# ── sheet readers ─────────────────────────────────────────────────────────────

def read_consolidated(wb):
    """
    New Consoliated — UK in-bond wines.
    Columns (after header row):
      0  Company | 1 Category | 2 Sub-Region | 3 Wine | 4 Vintage | 5 Amount
      6  Cost    | 7 (blank)  | 8 Deliver    | 9 For Sale | 10 Sold
      11 Asking price | 12 Remaining
    """
    ws = wb["New Consoliated"]
    rows = list(ws.iter_rows(values_only=True))

    header_idx = next(i for i, r in enumerate(rows) if r[0] == "Company")
    data_rows  = rows[header_idx + 1:]

    wines, listings = [], []

    for row in data_rows:
        name     = row[3]
        merchant = row[0]

        if not name or not merchant:
            continue

        bottle_count = parse_bottle_count(row[5])
        if not bottle_count:
            continue

        cost_total    = safe_float(row[6])
        for_sale_flag = row[9] and str(row[9]).strip().lower() == "x"
        asking_total  = safe_float(row[11])
        remaining     = safe_float(row[12])

        # Skip lots fully delivered/sold and NOT listed (they appear in Wien/Vinotheque tabs)
        if remaining == 0 and not for_sale_flag:
            continue

        merchant = normalise_merchant(merchant)
        cpb      = cost_per_bottle(cost_total, bottle_count)
        status   = "listed" if for_sale_flag else "in_storage"

        wine = {
            "name":             str(name).strip(),
            "category":         str(row[1]).strip() if row[1] else None,
            "sub_region":       str(row[2]).strip() if row[2] else None,
            "vintage":          safe_vintage(row[4]),
            "merchant":         merchant,
            "storage_location": storage_for_merchant(merchant),
            "bottle_count":     bottle_count,
            "bottle_format":    "75cl",
            "cost_per_bottle":  cpb,
            "status":           status,
            "paid":             True,
        }

        # For listed wines, record asking price and create a sell_listing entry
        if for_sale_flag and asking_total and bottle_count:
            asking_ppb = round(asking_total / bottle_count, 2)
            wine["value_per_bottle"] = asking_ppb
            listings.append({
                "_wine_name_key": f"{name}|{row[4]}|{merchant}",
                "merchant":       merchant,
                "bottles_listed": bottle_count,
                "list_price_per_bottle": asking_ppb,
                "status":         "active",
            })

        wines.append(wine)

    return wines, listings


def read_vinotheque(wb):
    """
    VINOTHEQUE — Vienna professional storage.
    Columns: 0 Delivery date | 1 Wine | 2 Region/Sub-region | 3 Vintage
             4 Amount | 5 Cost | 6 Invoice no | 7 From (original merchant) | 8 Notes
    """
    ws = wb["VINOTHEQUE"]
    rows = list(ws.iter_rows(values_only=True))

    header_idx = next(i for i, r in enumerate(rows) if r[1] == "Wine")
    data_rows  = rows[header_idx + 1:]

    wines = []
    for row in data_rows:
        name = row[1]
        if not name:
            continue

        bottle_count = parse_bottle_count(row[4])
        if not bottle_count:
            continue

        cost_total = safe_float(row[5])
        merchant   = normalise_merchant(row[7])
        cpb        = cost_per_bottle(cost_total, bottle_count)

        wines.append({
            "name":             str(name).strip(),
            "sub_region":       str(row[2]).strip() if row[2] else None,
            "vintage":          safe_vintage(row[3]),
            "merchant":         merchant,
            "storage_location": "Vinotheque",
            "bottle_count":     bottle_count,
            "bottle_format":    "75cl",
            "cost_per_bottle":  cpb,
            "invoice_no":       str(row[6]).strip() if row[6] else None,
            "notes":            str(row[8]).strip() if row[8] else None,
            "status":           "in_storage",
            "paid":             True,
        })

    return wines


def read_wien(wb):
    """
    Wien — Vienna home cellar.
    Columns: 0 Order | 1 Region | 3 Sub-Region | 5 Wine | 9 Vintage
             10 Amount | 11 Category | 14 From (original merchant)
    Note: no cost data in Wien tab.
    """
    ws = wb["Wien"]
    rows = list(ws.iter_rows(values_only=True))

    header_idx = next(i for i, r in enumerate(rows) if r[0] == "Order")
    data_rows  = rows[header_idx + 1:]

    wines = []
    for row in data_rows:
        order = row[0]
        name  = row[5]
        if not name or not order:
            continue

        bottle_count = parse_bottle_count(row[10])
        if not bottle_count:
            continue

        merchant = normalise_merchant(row[14])

        wines.append({
            "name":             str(name).strip(),
            "category":         str(row[11]).strip() if row[11] else None,
            "sub_region":       str(row[3]).strip() if row[3] else None,
            "vintage":          safe_vintage(row[9]),
            "merchant":         merchant,
            "storage_location": "Wien",
            "bottle_count":     bottle_count,
            "bottle_format":    "75cl",
            "status":           "in_storage",
            "paid":             True,
        })

    return wines


# ── insert helpers ────────────────────────────────────────────────────────────

def insert_batch(table, records, batch_size=100):
    inserted = 0
    for i in range(0, len(records), batch_size):
        batch = records[i : i + batch_size]
        supabase.table(table).insert(batch).execute()
        inserted += len(batch)
        print(f"    {inserted}/{len(records)}", end="\r")
    print()
    return inserted


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    print(f"Opening {EXCEL_FILE} …")
    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

    print("Reading New Consoliated (UK in-bond) …")
    uk_wines, listings = read_consolidated(wb)
    print(f"  {len(uk_wines)} lots  |  {len(listings)} listed for sale")

    print("Reading VINOTHEQUE (Vienna professional storage) …")
    vino_wines = read_vinotheque(wb)
    print(f"  {len(vino_wines)} lots")

    print("Reading Wien (Vienna home cellar) …")
    wien_wines = read_wien(wb)
    print(f"  {len(wien_wines)} lots")

    wb.close()

    all_wines = uk_wines + vino_wines + wien_wines
    print(f"\nTotal: {len(all_wines)} lots across all locations")
    print()

    # Confirm before inserting
    answer = input("Insert into Supabase? (yes/no): ").strip().lower()
    if answer != "yes":
        print("Aborted.")
        return

    print("\nInserting wines …")
    insert_batch("wines", all_wines)

    # Link sell_listings to inserted wine IDs
    if listings:
        print("\nLinking sell listings …")
        result = supabase.table("wines").select("id,name,vintage,merchant").eq("status", "listed").execute()
        wine_map = {
            f"{w['name']}|{w['vintage']}|{w['merchant']}": w["id"]
            for w in result.data
        }
        linked = []
        for listing in listings:
            key     = listing.pop("_wine_name_key")
            wine_id = wine_map.get(key)
            if wine_id:
                listing["wine_id"] = wine_id
                linked.append(listing)

        if linked:
            insert_batch("sell_listings", linked)
            print(f"  {len(linked)} sell listings created")
        else:
            print("  Could not match any listings to wine IDs — check manually")

    print("\nDone.")


if __name__ == "__main__":
    main()
